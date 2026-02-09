import pandas as pd
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================= CONFIG =================

entrada="reporte-nota-2025.xlsx"
salida="resumen_calificaciones_2025.xlsx"

# ================= FUNCIONES =================

def limpiar(txt):
    txt=str(txt)
    txt=unicodedata.normalize('NFKD',txt).encode('ascii','ignore').decode('ascii')
    return txt.lower().replace(" ","")

# ================= LEER EXCEL =================

df=pd.read_excel(entrada)

original=list(df.columns)

cols={limpiar(c):c for c in original}

def buscar(nombre):
    nombre=limpiar(nombre)
    for k,v in cols.items():
        if nombre in k:
            return v
    raise Exception(f"No encuentro columna {nombre}\nColumnas reales:\n{original}")

sexo=buscar("sexo")
rango=buscar("rango")
grado=buscar("grado")
area=buscar("area")

df[sexo]=df[sexo].astype(str)
df[rango]=df[rango].astype(str)

df["R"]=df[rango].str[:2]

rangos=["R1","R2","R3","R4"]

# ================= HOJA 1 =================

h1=[]

for r in rangos:
    f=len(df[(df.R==r)&(df[sexo].str.lower().str.contains("muj"))])
    m=len(df[(df.R==r)&(~df[sexo].str.lower().str.contains("muj"))])
    h1.append([r,f,m,f+m])

df1=pd.DataFrame(h1,columns=["Rango","Femenino","Masculino","Total"])

total=df1.Total.sum()
df1["Porcentaje"]=df1.Total/total
df1.loc["TOTAL"]=["TOTAL",df1.Femenino.sum(),df1.Masculino.sum(),total,1]

# ================= HOJA 2 =================

rows=[]
for n in df[grado].dropna().unique():
    for r in rangos:
        f=len(df[(df[grado]==n)&(df.R==r)&(df[sexo].str.lower().str.contains("muj"))])
        m=len(df[(df[grado]==n)&(df.R==r)&(~df[sexo].str.lower().str.contains("muj"))])
        rows.append([n,r,f,m,f+m])

df2=pd.DataFrame(rows,columns=["Nivel","Rango","Femenino","Masculino","Total"])

# ================= HOJA 3 =================

rows=[]
for a in df[area].dropna().unique():
    for r in rangos:
        f=len(df[(df[area]==a)&(df.R==r)&(df[sexo].str.lower().str.contains("muj"))])
        m=len(df[(df[area]==a)&(df.R==r)&(~df[sexo].str.lower().str.contains("muj"))])
        rows.append([a,r,f,m,f+m])

df3=pd.DataFrame(rows,columns=["Area","Rango","Femenino","Masculino","Total"])

# ================= EXPORTAR =================

with pd.ExcelWriter(salida,engine="openpyxl") as w:
    df1.to_excel(w,sheet_name="1_Rango",index=False)
    df2.to_excel(w,sheet_name="2_Nivel",index=False)
    df3.to_excel(w,sheet_name="3_Area",index=False)

# ================= FORMATO =================

wb=load_workbook(salida)

azul=PatternFill("solid",fgColor="00B0F0")
blanco=Font(color="FFFFFF",bold=True)
centro=Alignment(horizontal="center",vertical="center")
borde=Border(left=Side(style="thin"),right=Side(style="thin"),
top=Side(style="thin"),bottom=Side(style="thin"))

for ws in wb:

    ws.insert_rows(1,3)
    max_col=ws.max_column
    letra=get_column_letter(max_col)

    ws.merge_cells(f"A1:{letra}1")
    ws.merge_cells(f"A2:{letra}2")

    ws["A1"]="REPORTE DE CALIFICACIONES 2025"
    ws["A2"]=ws.title.replace("_"," ")

    ws["A1"].font=Font(bold=True,size=16,color="FFFFFF")
    ws["A2"].font=Font(bold=True,size=12,color="FFFFFF")

    for c in range(1,max_col+1):
        ws.cell(1,c).fill=azul
        ws.cell(2,c).fill=azul

    for c in ws[4]:
        c.fill=azul
        c.font=blanco

    for fila in ws.iter_rows(min_row=4):
        for celda in fila:
            celda.border=borde
            celda.alignment=centro

    last=ws.max_row
    for c in range(1,max_col+1):
        ws.cell(last,c).font=Font(bold=True)

ws1=wb["1_Rango"]
for i in range(5,ws1.max_row):
    ws1[f"E{i}"].number_format="0%"

wb.save(salida)

print("EXCEL CREADO:",salida)
